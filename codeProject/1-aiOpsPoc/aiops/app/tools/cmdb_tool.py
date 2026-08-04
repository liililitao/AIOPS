"""
CMDB 查询工具 - Agent 可调用的 Tool

当前实现: 读取 xlsx 文件，两级匹配查询
- 第一优先级: 精确匹配 Resource Name (告警中的 id 字段)
- 第二优先级: 模糊匹配域名 (告警中的 properties_hostname)

后期可改为 API 调用，只需修改此文件即可。
"""

import logging
from typing import Optional

import openpyxl
from langchain.tools import tool

from app.config import get_settings
from app.schemas.cmdb import CmdbLookupResult

logger = logging.getLogger("aiops.cmdb")

# CMDB 数据缓存 (避免每次查询都重新读取 xlsx)
_cache: Optional[dict] = None


def _load_cmdb() -> dict:
    """加载 CMDB 数据到内存缓存"""
    global _cache
    if _cache is not None:
        return _cache

    settings = get_settings()
    cache = {
        "iaas_rows": [],       # Azure IaaS 行数据 [{col: val}, ...]
        "paas_rows": [],       # Azure PaaS 行数据
        "system_rows": [],     # Computer System List 行数据
        "iaas_headers": [],
        "paas_headers": [],
        "system_headers": [],
    }

    xlsx_path = settings.cmdb_xlsx_path
    if not xlsx_path:
        logger.warning("CMDB xlsx 路径未配置或文件不存在")
        _cache = cache
        return cache

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        # 加载 Azure IaaS
        if "Azure IaaS" in wb.sheetnames:
            ws = wb["Azure IaaS"]
            headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            cache["iaas_headers"] = headers
            for row in ws.iter_rows(min_row=2, values_only=True):
                cache["iaas_rows"].append(dict(zip(headers, [str(v or "") for v in row])))

        # 加载 Azure PaaS
        if "Azure PaaS" in wb.sheetnames:
            ws = wb["Azure PaaS"]
            headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            cache["paas_headers"] = headers
            for row in ws.iter_rows(min_row=2, values_only=True):
                cache["paas_rows"].append(dict(zip(headers, [str(v or "") for v in row])))

        # 加载 Computer System List
        if "Computer System List" in wb.sheetnames:
            ws = wb["Computer System List"]
            headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            cache["system_headers"] = headers
            for row in ws.iter_rows(min_row=2, values_only=True):
                cache["system_rows"].append(dict(zip(headers, [str(v or "") for v in row])))

        wb.close()
        logger.info(
            f"CMDB 加载完成: IaaS={len(cache['iaas_rows'])}, "
            f"PaaS={len(cache['paas_rows'])}, "
            f"Systems={len(cache['system_rows'])}"
        )
    except Exception as e:
        logger.error(f"[CMDB] xlsx load failed: {e}")

    _cache = cache
    return cache


def _lookup_by_resource_id(resource_id: str) -> Optional[CmdbLookupResult]:
    """第一优先级: 精确匹配 Resource Name"""
    cmdb = _load_cmdb()

    # 在 Azure PaaS 中搜索
    for i, row in enumerate(cmdb["paas_rows"]):
        if row.get("Resource Name", "").strip() == resource_id.strip():
            return CmdbLookupResult(
                found=True,
                match_type="exact",
                resource_name=row.get("Resource Name", ""),
                resource_type=row.get("Resource Type", ""),
                environment=row.get("Environment", "Unknown"),
                subscription=row.get("SUBSCRIPTION", ""),
                source_sheet="Azure PaaS",
                source_row=i + 2,
            )

    # 在 Azure IaaS 中搜索
    for i, row in enumerate(cmdb["iaas_rows"]):
        if row.get("Server Name", "").strip() == resource_id.strip():
            return CmdbLookupResult(
                found=True,
                match_type="exact",
                resource_name=row.get("Server Name", ""),
                resource_type=row.get("VM", ""),
                environment=row.get("Environment", "Unknown"),
                subscription=row.get("Subscription Name", ""),
                source_sheet="Azure IaaS",
                source_row=i + 2,
            )

    return None


def _lookup_by_hostname(hostname: str) -> Optional[CmdbLookupResult]:
    """第二优先级: 模糊匹配域名（在 Computer System List 的 域名和证书 列）"""
    cmdb = _load_cmdb()

    if not hostname:
        return None

    hostname_lower = hostname.strip().lower()

    for i, row in enumerate(cmdb["system_rows"]):
        domains = row.get("域名和证书", "")
        if not domains:
            continue
        # 域名列可能包含多个域名，用换行或空格分隔
        domain_list = [d.strip().lower() for d in domains.replace("\n", " ").split() if d.strip()]
        for d in domain_list:
            if hostname_lower in d or d in hostname_lower:
                # 获取订阅名称，进一步在 PaaS/IaaS 中查找环境
                subscription = row.get("订阅名称", "")
                env = _find_env_by_subscription(subscription)
                return CmdbLookupResult(
                    found=True,
                    match_type="fuzzy",
                    resource_name=row.get("系统名称 (System Name)", ""),
                    resource_type="",
                    environment=env,
                    subscription=subscription,
                    source_sheet="Computer System List",
                    source_row=i + 2,
                )

    return None


def _find_env_by_subscription(subscription: str) -> str:
    """根据订阅名称推断环境"""
    if not subscription:
        return "Unknown"
    sub_upper = subscription.upper()
    # 订阅名称中包含 PRD 或 PROD → Production
    if "PRD" in sub_upper or "PROD" in sub_upper:
        return "Production"
    # 包含 TST 或 DEV → Non-Production
    if "TST" in sub_upper or "DEV" in sub_upper:
        return "Non-Production"
    return "Unknown"


def invalidate_cache():
    """清除 CMDB 缓存（xlsx 文件更新后调用）"""
    global _cache
    _cache = None
    logger.info("CMDB 缓存已清除")


@tool
def cmdb_lookup(resource_id: str = "", hostname: str = "") -> str:
    """
    查询 CMDB 数据库，根据 Azure 资源 ID 或域名获取设备所属环境（Production/Non-Production）及相关资产信息。

    输入参数:
    - resource_id: 告警中的 id 字段 (Azure 资源 ID)，如 "AGW-DAP-PRD-N3-01"
    - hostname: 告警中的 properties_hostname 字段 (域名)，如 "purview.novonordiskchina.com.cn"

    返回: JSON 格式的查询结果，包含 environment (Production/Non-Production/Unknown) 和匹配方式。
    """
    import json

    # 第一优先级: 精确匹配
    if resource_id:
        result = _lookup_by_resource_id(resource_id.strip())
        if result and result.found:
            return json.dumps(result.model_dump(), ensure_ascii=False, indent=2)

    # 第二优先级: 模糊匹配
    if hostname:
        result = _lookup_by_hostname(hostname.strip())
        if result and result.found:
            return json.dumps(result.model_dump(), ensure_ascii=False, indent=2)

    # 都没匹配到
    # 尝试从 resource_id 中推断（如 ID 包含 PRD 字样）
    env = "Unknown"
    if resource_id:
        rid_upper = resource_id.upper()
        if "PRD" in rid_upper or "PROD" in rid_upper:
            env = "Production"
        elif "TST" in rid_upper or "DEV" in rid_upper:
            env = "Non-Production"

    return json.dumps(
        CmdbLookupResult(
            found=False,
            match_type="none",
            environment=env,
            error=f"未在 CMDB 中找到匹配记录 (resource_id={resource_id}, hostname={hostname})",
        ).model_dump(),
        ensure_ascii=False,
        indent=2,
    )
